"""Deterministic loader for the assessment workbook (source of truth).

Returns typed row dicts (datetimes parsed, booleans preserved, None kept as
None — unknown fault attribution is meaningful to the domain layer). Used by
both the seeder and the Layer A tests so there is exactly one interpretation
of the workbook. No rows are invented, inferred, or dropped.
"""

from pathlib import Path

import openpyxl

from backend.domain.timebase import parse_ts

WORKBOOK_NAME = "ParcelPilot_Assessment_Data.xlsx"

_ORDER_TS_FIELDS = (
    "booked_at",
    "pickup_window_start",
    "pickup_window_end",
    "pickup_actual_at",
    "cancellation_requested_at",
)


def _sheet_rows(workbook, sheet_name):
    ws = workbook[sheet_name]
    header = [cell.value for cell in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw[0] is None:
            continue
        rows.append(dict(zip(header, raw)))
    return rows


def _normalize_order(row):
    out = dict(row)
    for field in _ORDER_TS_FIELDS:
        out[field] = parse_ts(out.get(field))
    out["shipment_fee_inr"] = int(out["shipment_fee_inr"])
    return out


def _normalize_ticket(row):
    out = dict(row)
    out["created_at"] = parse_ts(out["created_at"])
    out["last_customer_message_at"] = parse_ts(out.get("last_customer_message_at"))
    return out


def load_dataset(data_pack_dir):
    """Load accounts/orders/tickets from the assessment workbook."""
    path = Path(data_pack_dir) / WORKBOOK_NAME
    workbook = openpyxl.load_workbook(path, data_only=True)
    return {
        "accounts": _sheet_rows(workbook, "accounts"),
        "orders": [_normalize_order(r) for r in _sheet_rows(workbook, "orders")],
        "tickets": [_normalize_ticket(r) for r in _sheet_rows(workbook, "tickets")],
    }
